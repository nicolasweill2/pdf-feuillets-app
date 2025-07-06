import fitz
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import matplotlib.pyplot as plt


def process_folder(folder_path: str) -> str:
    # Liste PDF triée
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
    pdf_files = sorted(pdf_files, key=lambda x: x[:2])  # Tri sur les 2 premiers caractères
    
    # --- Extraction nb pages + dimensions ---
    data = []
    max_pages = 0
    for pdf_file in pdf_files:
        pdf_path = os.path.join(folder_path, pdf_file)
        try:
            doc = fitz.open(pdf_path)
            row = {
                "nom fichier": pdf_file,
                "nb pages": len(doc)
            }
            for i, page in enumerate(doc):
                rect = page.rect
                width = round(rect.width * 25.4 / 72, 2)
                height = round(rect.height * 25.4 / 72, 2)
                row[f"page {i+1}"] = f"{width} mm × {height} mm"
            max_pages = max(max_pages, len(doc))
            data.append(row)
        except Exception as e:
            data.append({
                "nom fichier": pdf_file,
                "nb pages": f"Erreur : {e}"
            })
    columns = ["nom fichier", "nb pages"] + [f"page {i+1}" for i in range(max_pages)]
    df = pd.DataFrame(data).reindex(columns=columns)

    # --- Analyse feuillets ---
    def dimensions_similaires(dim1, dim2, tol=0):
        w1, h1 = dim1
        w2, h2 = dim2
        return (abs(w1 - w2) <= tol and abs(h1 - h2) <= tol) or \
               (abs(w1 - h2) <= tol and abs(h1 - w2) <= tol)

    feuillets_data = []
    for pdf_file in pdf_files:
        pdf_path = os.path.join(folder_path, pdf_file)
        try:
            doc = fitz.open(pdf_path)
            page_dims = []
            for page in doc:
                rect = page.rect
                width = round(rect.width * 25.4 / 72, 2)
                height = round(rect.height * 25.4 / 72, 2)
                page_dims.append((width, height))
            i = 0
            feuillets = []
            while i < len(page_dims) - 1:
                if dimensions_similaires(page_dims[i], page_dims[i + 1]):
                    feuillets.append((f"page {i + 1}", f"page {i + 2}", page_dims[i]))
                    i += 2
                else:
                    feuillets.append((f"page {i + 1}", None, page_dims[i]))
                    i += 1
            if i == len(page_dims) - 1:
                feuillets.append((f"page {i + 1}", None, page_dims[i]))
            for f in feuillets:
                page_A, page_B, dim = f
                format_label = f"{round(dim[0])}×{round(dim[1])} mm"
                feuillets_data.append({
                    "pdf": pdf_file,
                    "page 1": page_A,
                    "page 2": page_B if page_B else "",
                    "format": format_label
                })
        except Exception as e:
            print(f"Erreur pour {pdf_file} : {e}")

    feuillets_df = pd.DataFrame(feuillets_data)
    feuillet_counts = feuillets_df["pdf"].value_counts().to_dict()
    feuillets_df["nombre de feuillets"] = feuillets_df["pdf"].map(feuillet_counts)

    # --- Sauvegarder feuilles nb page + nb feuillet ---
    output_excel = os.path.join(folder_path, f"pages_colonnes_dossier_{os.path.basename(folder_path)}.xlsx")

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="nb page", index=False)

    with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        feuillets_df.to_excel(writer, sheet_name="nb feuillet", index=False)

    # --- Ajuster largeur colonnes des 2 feuilles ---
    wb = load_workbook(output_excel)
    sheets_to_adjust = ["nb page", "nb feuillet"]
    for sheet_name in sheets_to_adjust:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(output_excel)

    # --- Générer le graphique de répartition ---
    format_counts = feuillets_df.groupby(["pdf", "format"]).size().reset_index(name="nb_feuillets")
    pivot_df = format_counts.pivot(index="pdf", columns="format", values="nb_feuillets").fillna(0)
    pivot_df = pivot_df.iloc[::-1]

    n_pdfs = pivot_df.shape[0]
    height_per_pdf = 0.4
    fig_width = 12
    fig_height = max(4, n_pdfs * height_per_pdf)

    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    pivot_df.plot(kind="barh", stacked=True, ax=ax)

    plt.subplots_adjust(left=0.25, right=0.85)
    ax.tick_params(axis='y', labelsize=8)
    plt.title("Répartition des formats de feuillets par fichier PDF")
    plt.xlabel("Nombre de feuillets")
    plt.legend(title="Format feuillet", bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.tight_layout()

    for container in ax.containers:
        labels = []
        for rect in container:
            width = rect.get_width()
            labels.append(f'{int(width)}' if width > 0 else '')
        ax.bar_label(container, labels=labels, label_type='center', fontsize=8)

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()

    # --- Préparer stats pour récapitulatif ---
    stats_pdf = pd.DataFrame({
        "nb_pages": df.groupby("nom fichier")["nb pages"].first(),
        "nb_feuillets": feuillets_df.groupby("pdf").size(),
        "nb_feuillets_recto_seul": feuillets_df[feuillets_df["page 2"] == ""].groupby("pdf").size(),
    })
    stats_pdf["nb_feuillets_orphelins"] = stats_pdf["nb_feuillets_orphelins"].fillna(0).astype(int)

    formats_par_pdf = format_counts.groupby("pdf").apply(
        lambda df: "; ".join(f"{row['format']} ({row['nb_feuillets']})" for _, row in df.iterrows())
    )
    stats_pdf["formats_counts"] = formats_par_pdf

    total_feuillets = len(feuillets_df)
    total_pages = df["nb pages"].sum()
    total_orphelins = (feuillets_df["page 2"] == "").sum()

    formats_counts_global = feuillets_df["format"].value_counts().to_frame().reset_index()
    formats_counts_global.columns = ["format", "effectif"]

    # --- Génération feuille "Récapitulatif" ---
    if "Récapitulatif" in wb.sheetnames:
        wb.remove(wb["Récapitulatif"])
    ws = wb.create_sheet("Récapitulatif")

    img = OpenpyxlImage(buffer)
    img.anchor = "A1"
    ws.add_image(img)

    start_row = 2
    start_col = 21
    recap_col = start_col + stats_pdf.shape[1] + 2
    format_col = recap_col

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)

    ws.cell(row=start_row - 1, column=start_col, value="Statistiques par PDF").font = title_font
    for r_idx, row in enumerate(dataframe_to_rows(stats_pdf.reset_index(), index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
            if r_idx == start_row:
                cell.font = header_font

    ws.cell(row=start_row - 1, column=recap_col, value="Résumé global").font = title_font
    resume_data = [
        ("Nombre total de pages", total_pages),
        ("Nombre total de feuillets", total_feuillets),
        ("Nombre total feuillets recto seul", total_orphelins),
    ]
    for i, (label, value) in enumerate(resume_data):
        label_cell = ws.cell(row=start_row + i, column=recap_col, value=label)
        value_cell = ws.cell(row=start_row + i, column=recap_col + 1, value=value)
        for cell in [label_cell, value_cell]:
            cell.border = thin_border
            cell.alignment = center_align

    ws.cell(row=start_row + len(resume_data) + 2, column=format_col, value="Effectif formats global").font = title_font
    for i, row in formats_counts_global.iterrows():
        format_cell = ws.cell(row=start_row + len(resume_data) + 3 + i, column=format_col, value=row["format"])
        count_cell = ws.cell(row=start_row + len(resume_data) + 3 + i, column=format_col + 1, value=row["effectif"])
        for cell in [format_cell, count_cell]:
            cell.border = thin_border
            cell.alignment = center_align

    for col_idx in range(start_col, format_col + 3):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_excel)

    print(f"Traitement terminé. Fichier Excel généré : {output_excel}")

    return output_excel
